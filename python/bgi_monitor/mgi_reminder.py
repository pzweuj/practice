# coding=utf-8
# pzw
# 20221112

import paramiko
import requests
import time
from datetime import datetime
import openpyxl

# 消息命令，使用markdown格式
def sender(key, msgStr):
    params = {
        "key": key
    }
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": msgStr
        }
    }
    requests.post("https://qyapi.weixin.qq.com/cgi-bin/webhook/send", params=params, json=data)

# 时间段警告
def TimeTagWarning(nowChipName, nowStep, nowTime):
    wb = openpyxl.load_workbook("TimeTag.xlsx")
    ws = wb.active
    now = datetime.now()
    toast = ""
    processDict = {
        "Begin": "Step0",
        "C039": "Step1",
        "C078": "Step2",
        "C117": "Step3",
        "Analysis": "Step4",
        "Done": "Step5"
    }
    timeProcess = {}
    n = 2
    while n >= 2:
        chipName = ws["A"+str(n)].value
        if chipName == None:
            break
        process = ws["B"+str(n)].value
        processTimeTag = ws["C"+str(n)].value
        timeProcess[chipName] = [process, processTimeTag, n]
        n += 1
    
    # 如果没有证明未记录过，新增一行
    if not nowChipName in timeProcess.keys():
        ws["A"+str(n)].value = nowChipName
        ws["B"+str(n)].value = nowStep
        ws["C"+str(n)].value = nowTime

    # 如果存在过，则需要判断状态
    else:
        # 状态相同，判断时间间隔多久
        if nowStep == timeProcess[nowChipName][0]:
            if nowStep in ["Begin", "C039", "C078"]:
                # 时间间隔大于3小时，可能出错
                if (now - timeProcess[nowChipName][1]).total_seconds() > (3 * 3600):
                    toast = "【提醒】芯片号：" + nowChipName + " 测序过程可能出现问题，请及时查看测序仪状态。"
                else:
                    pass
            elif nowStep in ["C117"]:
                # 从下机完成到分析结果生成大概1个小时
                if (now - timeProcess[nowChipName][1]).total_seconds() > 3600:
                    toast = "【提醒】芯片号：" + nowChipName + " 数据分析过程可能出现问题，请及时查看服务器后台状态。"
                else:
                    pass
            elif nowStep in ["Done"]:
                pass
            elif nowStep in ["unknown"]:
                pass
            else:
                pass
        
        # 状态不同，更新状态及时间戳
        else:
            colNum = timeProcess[nowChipName][2]
            ws["B"+str(colNum)].value = nowStep
            ws["C"+str(colNum)].value = nowTime
    
    wb.save("TimeTag.xlsx")
    wb.close()
    return toast

# 监控下机
def sequenceMonitor(hostname, username, password, path):
    port = 22
    sf = paramiko.Transport((hostname, port))
    sf.connect(username=username, password=password)
    sftp = paramiko.SFTPClient.from_transport(sf)

    # 已完成的列表
    ## 以是否存在117cycle作为完成标记
    sequenceFinishedList = []
    with open("sequenceFinished.txt", "r", encoding="utf-8") as f:
        for line in f:
            sequenceFinishedList.append(line.replace("\n", ""))
    # print(sequenceFinishedList)
    # 监控以下目录
    sftp.chdir(path)
    pathList = sftp.listdir(path)
    chipName = "Unknown"
    toast = ""
    nowStatus = "unknown"
    for p in pathList:
        # 在已完成列表
        try:
            chipName = p.split("_")[1]
        except:
            pass
        # print(chipName)
        if chipName in sequenceFinishedList:
            continue
        # 不在，则检测完成到哪一步
        else:
            # 尝试打开状态文件
            statusDict = {}
            try:
                status = open("status.txt", "r", encoding="utf-8")
                for line in status:
                    lines = line.replace("\n", "").split("\t")
                    statusDict[lines[0]] = lines[1]
                status.close()
            except:
                pass
            newPath = path + "/" + p + "/die1"
            try:
                sftp.chdir(newPath)
                chipName = p.split("_")[1]
            # 转不过去就可能是测序未开始又或者不是文件夹
            except:
                continue
            subList = sftp.listdir(newPath)
            # 已完成
            if "C117" in subList:
                toast = "【测序】芯片号：" + chipName + " 测序运行进度100%，预计正在拆分，拆分完成后将进行数据上传。稍后请检查分析状态。"
                finishFile = open("sequenceFinished.txt", "a", encoding="utf-8")
                finishFile.write(chipName + "\n")
                nowStatus = "C117"
            elif "C078" in subList:
                if chipName in statusDict.keys():
                    if statusDict[chipName] != "C078":
                        toast = "【测序】芯片号：" + chipName + " 测序运行进度66%"
                        status = open("status.txt", "w", encoding="utf-8")
                        status.write(chipName + "\tC078\n")
                        status.close()
                        nowStatus = "C078"
                    else:
                        continue
                else:
                    toast = "【测序】芯片号：" + chipName + " 测序运行进度66%"
                    status = open("status.txt", "w", encoding="utf-8")
                    status.write(chipName + "\tC078\n")
                    status.close()
                    nowStatus = "C078"
            elif "C039" in subList:
                if chipName in statusDict.keys():
                    if statusDict[chipName] != "C039":
                        toast = "【测序】芯片号：" + chipName + " 测序运行进度33%"
                        status = open("status.txt", "w", encoding="utf-8")
                        status.write(chipName + "\tC039\n")
                        status.close()
                        nowStatus = "C039"
                    else:
                        continue
                else:
                    toast = "【测序】芯片号：" + chipName + " 测序运行进度33%"
                    status = open("status.txt", "w", encoding="utf-8")
                    status.write(chipName + "\tC039\n")
                    status.close()
                    nowStatus = "C039"
            else:
                nowStatus = "Begin"
                continue
    
    # 关闭sftp连接
    sftp.close()
    return [toast, nowStatus, chipName]

# 监控分析
## 在服务器能保持外网连接后，不需要使用sftp，重写这个模块
def analysisMonitor(hostname, username, password, path):
    port = 22
    sf = paramiko.Transport((hostname, port))
    sf.connect(username=username, password=password)
    sftp = paramiko.SFTPClient.from_transport(sf)

    # 已完成列表
    analysisFinishList = []
    analysisRunList = []
    with open("analysisFinished.txt", "r", encoding="utf-8") as f:
        for line in f:
            lines = line.replace("\n", "").split("\t")
            analysisFinishList.append(lines[0])
            analysisRunList.append(lines[1])
    analysisFinishList = list(set(analysisFinishList))

    # 监控以下目录
    ## /data/pathonge/data/result
    sftp.chdir(path)
    pathList = sftp.listdir(path)
    toast = ""
    nowStatus = "unknown"
    # p是批次号
    for p in pathList:
        # 已在列表中
        try:
            chipName = p.split("_")[1]
        except:
            chipName = "-"
        if p in analysisFinishList:
            continue
        else:
            # 查看是否完成
            projectPath = path + "/" + p
            sftp.chdir(projectPath)
            projectPathList = sftp.listdir(projectPath)
            # pj是任务号
            for pj in projectPathList:
                if pj in analysisRunList:
                    continue
                projectSubPath = projectPath + "/" + pj
                projectSubPathList = sftp.listdir(projectSubPath)
                for d in projectSubPathList:
                    sftp.chdir(projectSubPath + "/" + d)
                    projectSubSonList = sftp.listdir(projectSubPath + "/" + d)
                    if "All.done" in projectSubSonList:
                        nowStatus = "Done"
                        toast = "【分析】芯片号：" + chipName + "\t任务号：" + pj + " 已分析完成，请及时导出并发送分析结果。同时请关注同一芯片是否有其他任务。"
                        analysisFinishFile = open("analysisFinished.txt", "a", encoding="utf-8")
                        analysisFinishFile.write(p + "\t" + pj + "\n")
                        analysisFinishFile.close()
    sftp.close()
    return [toast, nowStatus, chipName]

# Main
while True:
    # 测序过程监控
    nowTime = datetime.now()
    now = nowTime.strftime("%Y-%m-%d %H:%M:%S")
    try:
        toastList = sequenceMonitor("192.168.1.2", "username", "password", "/test/path")
        chipName = toastList[2]
        toast = TimeTagWarning(chipName, toastList[1], nowTime)
        if toast == "":
            # 这个优先
            toast = toastList[0]
    except:
        toast = "error"
    logFile = open("running.log", "a", encoding="utf-8")
    if toast == "error":
        toast = "【失败】测序监控执行失败，可能是网络断连\t"
        logFile.write(toast + now + "\n")
    elif toast != "":
        logFile.write(toast + "\t" + now + "\n")
        # 推送机器
        sender("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa", toast)
    else:
        toast = "【正常】测序监控正常执行\t"
        logFile.write(toast + now + "\n")
    logFile.close()

    time.sleep(30)

    # 分析流程监控

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        toastList = analysisMonitor("192.168.1.1", "username", "password", "/data/result")
        chipName = toastList[2]
        toast = TimeTagWarning(chipName, toastList[1], nowTime)
        if toast == "":
            toast = toastList[0]
    except:
        toast = "error"
    logFile = open("running.log", "a", encoding="utf-8")
    if toast == "error":
        toast = "【失败】分析监控执行失败，可能是网络断连\t"
        logFile.write(toast + now + "\n")
    elif toast != "":
        logFile.write(toast + "\t" + now + "\n")
        # 推送机器
        sender("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa", toast)
    else:
        toast = "【正常】分析监控正常执行\t"
        logFile.write(toast + now + "\n")
    logFile.close()
    time.sleep(10*60)

# end

