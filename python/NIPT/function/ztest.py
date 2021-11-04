# coding=utf-8
# pzw
# NIPT analysis
# 20211101
# 依赖于bwa, samtools, sambamba, bedtools, R

import os
import sys
import pandas as pd
import shutil
import argparse
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import load_workbook

# 获得程序运行路径
def getAbsPath():
    now = os.path.abspath(os.path.dirname(sys.argv[0]))
    return now

# 获得Contig大小，只需要常染色体chr1-chr22
# reference需要预先建立索引
# samtools faidx reference.fa
def GetFastaBinsGCBias(reference, tmpDir, output):
    if not os.path.exists(tmpDir):
        os.makedirs(tmpDir)
    fai = open(reference + ".fai", "r")
    contig = open(tmpDir + "/contig.size", "w")
    contigList = ["chr" + str(x) for x in range(1, 23)]
    for line in fai:
        lines = line.split("\t")
        chrom = lines[0]
        length = lines[1]
        if chrom in contigList:
            contig.write(chrom + "\t" + length + "\n")
    contig.close()
    fai.close()
    cmd = """
        bedtools makewindows -g {tmpDir}/contig.size -w 300000 > {tmpDir}/contig.300kb.bed
        bedtools nuc -fi {reference} -bed {tmpDir}/contig.300kb.bed | cut -f 1-3,5 > {output}
    """.format(reference=reference, tmpDir=tmpDir, output=output)
    os.system(cmd)

# 获得每个bam文件的每个bin的counts数
# bam文件来源
# bwa mem
# samtools sort
# sambamba markdup -r
def GetBamCounts(bamPath, gcBed, output):
    if not os.path.exists(output):
        os.makedirs(output)
    for bam in os.listdir(bamPath):
        if bam.endswith(".bam"):
            name = bam.replace(".bam", ".counts")
            bamFile = bamPath + "/" + bam
            cmd = """
                bedtools coverage -a {gcBed} -b {bamFile} | cut -f 1-5 \\
                    | sed '1i chr\\tstart\\tend\\tGC\\tRC' > {output}/{name}
            """.format(gcBed=gcBed, bamFile=bamFile, output=output, name=name)
            os.system(cmd)

# 适用GC比例校正read counts
def LoessCorrectCountsWithGC(countPath, output):
    if not os.path.exists(output):
        os.makedirs(output)
    script = getAbsPath() + "/gcContent.R"
    for count in os.listdir(countPath):
        gcCount = count.replace(".count", ".cor.count")
        cmd = """
            Rscript {script} {countPath}/{count} {output}/{gcCount}
        """.format(script=script, countPath=countPath, count=count, output=output, gcCount=gcCount)
        os.system(cmd)

# 形成Counts矩阵
def CountsMatrix(corPath, matrixFile):
    countDict = {}
    for txt in os.listdir(corPath):
        name = txt.replace(".cor.count", "")
        countDict[name] = {}
        f = open(corPath + "/" + txt, "r")
        for line in f:
            if not line.startswith("chr\tstart"):
                lines = line.replace("\n", "").split("\t")
                chromosome = lines[0]
                readCounts = lines[4]
                try:
                    countDict[name][chromosome] += float(readCounts)
                except:
                    countDict[name][chromosome] = 0
                    countDict[name][chromosome] += float(readCounts)
        f.close()
    matrix = open(matrixFile, "w")
    kList = list(countDict.keys())
    matrix.write("#SampleName\t" + "\t".join(kList) + "\n")
    for i in range(22):
        chr = "chr" + str(i+1)
        countList = []
        for k in kList:
            countList.append("%.2f" % countDict[k][chr])
        matrix.write(chr + "\t" + "\t".join(countList) + "\n")
    matrix.close()

# 计算Z值
def calZ(countsMatrix, reference, xlsx):
    ref = open(reference, "r", encoding="utf-8")
    chromDict = {}
    for line in ref:
        if not line.startswith("#"):
            lines = line.replace("\n", "").split("\t")
            chromDict[lines[0]] = [float(lines[1]), float(lines[2])]
    ref.close()
    
    df = pd.read_table(countsMatrix, sep="\t", header=0, index_col=0)
    colSumDict = {}
    for col in df.columns.values.tolist():
        colSumDict[col] = df[col].sum()
    df_UR = df
    for row in df_UR.index:
        for col in colSumDict.keys():
            df_UR.loc[row, col] = df_UR.loc[row, col] / colSumDict[col]
    df_Z = df_UR
    for row in df_Z.index:
        for col in df_Z.columns.values.tolist():
            df_Z.loc[row, col] = (df_UR.loc[row, col] - chromDict[row][0]) / chromDict[row][1]
    df_Z.to_excel(xlsx, sheet_name="Zscore", na_rep="NA")

# fix
def fixExcel(inputXlsx, outputXlsx):
    wb = load_workbook(inputXlsx)
    ws = wb.active
    colNum = ws.max_column
    n = 2
    while n <= colNum:
        for i in range(2, 24):
            if ws.cell(row=i, column=n).value < -3 or ws.cell(row=i, column=n).value > 3:
                ws.cell(row=i, column=n).fill = PatternFill(fgColor="EA9CAC", fill_type="solid")
                ws.cell(row=i, column=n).font = Font(color="922B21")
        n += 1
    wb.save(outputXlsx)

def main(reference, bamPath, refSet, countsFile, xlsx):
    tmpdir = getAbsPath() + "/tmp"
    GetFastaBinsGCBias(reference, tmpdir, tmpdir + "/gc.bed")
    GetBamCounts(bamPath, tmpdir + "/gc.bed", tmpdir + "/referenceCounts")
    LoessCorrectCountsWithGC(tmpdir + "/referenceCounts", tmpdir + "/referenceCorCounts")
    CountsMatrix(tmpdir + "/referenceCorCounts", countsFile)
    calZ(countsFile, refSet, tmpdir + "/tmp.xlsx")
    fixExcel(tmpdir + "/tmp.xlsx", xlsx)
    if os.path.exists(tmpdir):
        shutil.rmtree(tmpdir)
    print("结果生成：", xlsx)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="make NIPT results @pzweuj",
        prog="ztest.py",
        usage="python3 ztest.py [-h] -f <ref.fa> -r <ref file> -b <bam files dir path> -o <excel file> -c <count table>",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("-v", "--version", action="version",
        version="Version 0.1 20211104")
    parser.add_argument("-f", "--fasta", type=str,
        help="reference.fasta(contain chr1-22)")
    parser.add_argument("-r", "--ref", type=str,
        help="ref file create by reference.py")
    parser.add_argument("-b", "--bamPath", type=str,
        help="reference bam files directory")
    parser.add_argument("-c", "--counts", type=str,
        help="read counts output")
    parser.add_argument("-o", "--output", type=str,
        help="z score results")

    if len(sys.argv[1:]) == 0:
        parser.print_help()
        parser.exit()
    args = parser.parse_args()
    main(reference=args.fasta, refSet=args.ref, bamPath=args.bamPath, countsFile=args.counts, xlsx=args.output)
