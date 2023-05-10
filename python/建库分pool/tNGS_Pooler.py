# coding=utf-8
# pzw
# 20221111
# tNGS pooling 自动化

import os
import sys
import openpyxl
import pandas as pd
import math
from openpyxl.styles.borders import Border, Side
import time
import locale
import PySimpleGUI as sg


# 获得程序运行路径
def getAbsPath():
    now = os.path.abspath(os.path.dirname(sys.argv[0]))
    return now


# 读入表格
def readTable(inputXlsx, tableIndex, Dfrom=1, Rfrom=1):
    labInfo = openpyxl.load_workbook(inputXlsx).active
    infoDict = {}
    infoList = []
    DNAList = []
    RNAList = []
    otherList = []
    
    # 确认文库浓度在哪一列
    # 没那么离谱吧，定个26就够了吧
    libCol = 13
    c = 10
    while c in range(10, 27):
        libColName = labInfo.cell(1, c).value
        if libColName == "文库Qubit(ng/ul)":
            libCol = c
        c += 1

    n = 2
    while n >= 2:
        indexNum = labInfo.cell(n, 1).value
        if indexNum == None:
            break
        
        # if indexNum.startswith("D"):
        #     DNAList.append(int(indexNum.replace("D", "")))
        
        # elif indexNum.startswith("R"):
        #     RNAList.append(int(indexNum.replace("R", "")))

        if "-D" in indexNum:
            DNAList.append(int(indexNum.split("-D")[1]))
        elif "-R" in indexNum:
            RNAList.append(int(indexNum.split("-R")[1]))
        else:
            otherList.append(indexNum)

        # libraryQubit = labInfo.cell(n, 13).value
        libraryQubit = labInfo.cell(n, libCol).value
        tableNum = tableIndex + "_" + indexNum
        if indexNum in infoList:
            print("【错误】 存在相同的序号名")
        
        # if indexNum.startswith("D"):
        #     infoDict[tableNum] = ["D"+str(Dfrom), libraryQubit]
        #     Dfrom += 1
        # elif indexNum.startswith("R"):
        #     infoDict[tableNum] = ["R"+str(Rfrom), libraryQubit]
        #     Rfrom += 1
        # else:
        #     infoDict[tableNum] = [tableNum, libraryQubit]

        if "-D" in indexNum:
            infoDict[tableNum] = ["D" + str(Dfrom), libraryQubit]
            Dfrom += 1
        elif "-R" in indexNum:
            infoDict[tableNum] = ["R" + str(Rfrom), libraryQubit]
            Rfrom += 1
        else:
            infoDict[tableNum] = [tableNum, libraryQubit]

        n += 1
    
    if len(DNAList) == 0:
        DNAListMax = 0
    else:
        DNAListMax = max(DNAList)
    if len(RNAList) == 0:
        RNAListMax = 0
    else:
        RNAListMax = max(RNAList)
    return [infoDict, DNAListMax, RNAListMax]

# 获得当天日期
def getTime():
    locale.setlocale(locale.LC_CTYPE, "chinese")
    now = time.strftime("日期：%Y年%m月%d日")
    return now




#################################################
def main(class1File, class2File, class1Sig, class2Sig, outputExcel):

    class1List = readTable(class1File, class1Sig)
    class1 = class1List[0]
    
    if class2File != "":
        class2 = readTable(class2File, class2Sig, class1List[1]+1, class1List[2]+1)[0]
        # print(class2)
        mergeDict = {**class1, **class2}
    else:
        mergeDict = class1

    if not ".xlsx" in outputExcel:
        outputExcel = outputExcel + ".xlsx"

    # 样式
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    greenFill = openpyxl.styles.PatternFill("solid", fgColor="c6efce")
    greenFont = openpyxl.styles.Font(u"宋体", size=11, bold=True, italic=False, strike=False, color="006100")
    # redFill = openpyxl.styles.PatternFill("solid", fgColor="ffc7ce")
    # redFont = openpyxl.styles.Font(u"宋体", size=11, bold=True, italic=False, strike=False, color="9c0006")
    yellowFill = openpyxl.styles.PatternFill("solid", fgColor="ffeb9c")
    yellowFont = openpyxl.styles.Font(u"宋体", size=11, bold=True, italic=False, strike=False, color="9c6500")

    # 将输入的xlsx提取关键信息转换为txt，再导入pandas
    tmpOut = open(getAbsPath() + "/Data/tmp.txt", "w", encoding="utf-8")
    tmpOut.write("序号\t样本编号\t文库浓度\n")
    for m in mergeDict:
        outputList = [m, mergeDict[m][0], str(mergeDict[m][1])]
        tmpOut.write("\t".join(outputList) + "\n")
    tmpOut.close()
    df = pd.read_csv(getAbsPath() + "/Data/tmp.txt", sep="\t", header=0)
    df["文库大小"] = 300
    df["数据量"] = 16 / len(df)
    df["摩尔浓度"] = df["文库浓度"] * 1515 / df["文库大小"]

    # 计算Pooling
    df_sort = df.sort_values(by="文库浓度", ascending=False)
    df_sort["理论Pool"] = "NA"
    NAIndex = df_sort[df_sort["理论Pool"] == "NA"].index.tolist()
    n = 1
    while len(NAIndex) != 0:
        firstNAIndex = NAIndex[0]
        firstMol = df_sort.loc[firstNAIndex, "摩尔浓度"]
        firstDataAmount = df_sort.loc[firstNAIndex, "数据量"]
        
        # 按摩尔浓度排序，当摩尔浓度除以数据量小于1，认为这个样本浓度太低，需要舍弃
        if (firstMol / firstDataAmount) <= 1:
            break
        
        # 计算校正参数
        firstStep = (math.ceil(firstMol / firstDataAmount) // 10 + 1) * 10
        for i in NAIndex:
            df_sort.loc[i, "pooling"] = firstStep / df_sort.loc[i, "摩尔浓度"] * df_sort.loc[i, "数据量"]
            df_sort.loc[i, "fixNum"] = firstStep
        
        # 所有满足当前校正参数的样本的都会被改写
        changingIndex = df_sort[df_sort["pooling"] <= 10].index.tolist()
        for c in changingIndex:
            if c in NAIndex:
                df_sort.loc[c, "理论Pool"] = "P" + str(n)
                NAIndex.remove(c)
        n += 1

    df_sort["质量"] = df_sort["pooling"] * df_sort["文库浓度"]

    # 总共有几个pool，要保持顺序
    poolListInput = df_sort["理论Pool"].tolist()
    poolList = list(sorted(set(poolListInput), key=poolListInput.index))
    if "NA" in poolList:
        poolNum = len(poolList) - 1
        poolList.remove("NA")
        poolList.insert(0, "NA")
    else:
        poolNum = len(poolList)
    # print(poolList)

    # 计算各个pool的理论值
    poolDict = {}
    for p in range(poolNum + 1):
        if p == 0:
            continue
        poolName = "P" + str(p)
        df_pool = df_sort[df_sort["理论Pool"] == poolName]
        dataAmountAll = df_pool["数据量"].sum()
        volAll = df_pool["pooling"].sum()
        weightAll = df_pool["质量"].sum()
        theoreticalFc = weightAll / volAll
        librarySize = 300
        molFc = theoreticalFc * 1515 / librarySize
        poolDict[poolName] = {"dataAmount": dataAmountAll, "vol": volAll, "weight": weightAll, "Fc": theoreticalFc, "molFc": molFc}

    # 校正数
    m = 1
    lastPool = poolDict["P" + str(poolNum)]
    lastFix = lastPool["molFc"] / lastPool["dataAmount"]
    lastCheck = 1 / lastFix * m
    while lastCheck <= 10:
        m += 1
        lastCheck = 1 / lastFix * m
    fixNum = m - 1

    poolDataAmountAll = 0
    poolVolAll = 0
    poolWeightAll = 0
    for p in range(poolNum + 1):
        if p == 0:
            continue
        poolName = "P" + str(p)
        poolDict["P"+str(p)]["pooling"] = fixNum / poolDict["P"+str(p)]["molFc"] * poolDict["P"+str(p)]["dataAmount"]
        poolDict["P"+str(p)]["poolWeight"] = poolDict["P"+str(p)]["pooling"] * poolDict["P"+str(p)]["Fc"]
        poolDataAmountAll += poolDict["P"+str(p)]["dataAmount"]
        poolVolAll += poolDict["P"+str(p)]["pooling"]
        poolWeightAll += poolDict["P"+str(p)]["poolWeight"]
    theoreticalPoolFc = poolWeightAll / poolVolAll

    # 修正方向
    df_sort = df_sort.sort_values(by="文库浓度", ascending=True)
    poolNumAssen = poolNum
    poolNumCount = 1
    while poolNumAssen != 0:
        df_sort.loc[df_sort["理论Pool"] == "P" + str(poolNumCount), "理论Pool"] = "P_tmp_" + str(poolNumAssen)
        poolDict["P_tmp_" + str(poolNumAssen)] = poolDict["P" + str(poolNumCount)]
        poolNumAssen = poolNumAssen - 1
        poolNumCount += 1
    df_sort["理论Pool"] = df_sort.loc[:, "理论Pool"].str.replace("_tmp_", "")
    keyList = list(poolDict.keys())
    for k in keyList:
        if "_tmp_" in k:
            poolDict[k.replace("_tmp_", "")] = poolDict[k]
            del poolDict[k]

    ############### mix表  ##################
    # print(df_sort)
    poolingTemplate = openpyxl.load_workbook(getAbsPath() + "/Data/Pooling_template.xlsx")
    mixSheet = poolingTemplate["mix"]
    mixSheet.cell(3, 1).value = getTime()
    n = 6

    # print(poolDict)

    poolingFixNum = []
    lastPool = ""
    fills = [yellowFill, greenFill]
    fonts = [yellowFont, greenFont]
    poolNumStartIndexOrd = 75
    
    # 为了sum
    for i in df_sort.index.tolist():
        mixSheet.cell(n, 1).value = df_sort.loc[i, "序号"]
        mixSheet.cell(n, 2).value = "".join(df_sort.loc[i, "序号"].split("_")[1:])
        mixSheet.cell(n, 3).value = df_sort.loc[i, "文库浓度"]
        mixSheet.cell(n, 4).value = df_sort.loc[i, "文库大小"]
        mixSheet.cell(n, 5).value = round(df_sort.loc[i, "数据量"], 2)
        # mixSheet.cell(n, 6).value = round(df_sort.loc[i, "摩尔浓度"], 4)
        mixSheet.cell(n, 6).value = "=C" + str(n) + "*1515/D" + str(n)
        mixSheet.cell(n, 7).value = df_sort.loc[i, "样本编号"]
        # mixSheet.cell(n, 8).value = round(df_sort.loc[i, "pooling"], 1)
        # mixSheet.cell(n, 10).value = round(df_sort.loc[i, "质量"], 4)
        mixSheet.cell(n, 10).value = "=C" + str(n) + "*H" + str(n)

        # 当前的pool
        rPool = df_sort.loc[i, "理论Pool"]

        mixSheet.cell(n, 1).border = thin_border
        mixSheet.cell(n, 2).border = thin_border
        mixSheet.cell(n, 3).border = thin_border
        mixSheet.cell(n, 4).border = thin_border
        mixSheet.cell(n, 5).border = thin_border
        mixSheet.cell(n, 6).border = thin_border
        mixSheet.cell(n, 7).border = thin_border
        mixSheet.cell(n, 8).border = thin_border
        mixSheet.cell(n, 9).border = thin_border
        mixSheet.cell(n, 10).border = thin_border
        # mixSheet.cell(n, 11).border = thin_border

        if rPool != lastPool:
            poolingFixNum.append(int(df_sort.loc[i, "fixNum"]))
            fills = fills[::-1]
            fonts = fonts[::-1]
            poolNumStartIndex = n
            if rPool == "NA":
                mixSheet.cell(n, 11).value = "NA"
                mixSheet.cell(n, 12).value = "NA"
                mixSheet.cell(n, 13).value = "NA"
                mixSheet.cell(n, 14).value = "NA"
                mixSheet.cell(n, 15).value = rPool
            
            else:
                # mixSheet.cell(n, 11).value = round(poolDict[rPool]["dataAmount"], 2)
                poolSampleAmount = len(df_sort[df_sort["理论Pool"] == rPool].index)
                mixSheet.cell(n, 11).value = "=sum(E" + str(poolNumStartIndex) + ":E" + str(poolNumStartIndex+poolSampleAmount-1) + ")"
                # mixSheet.cell(n, 12).value = round(poolDict[rPool]["vol"], 1)
                mixSheet.cell(n, 12).value = "=sum(H" + str(poolNumStartIndex) + ":H" + str(poolNumStartIndex+poolSampleAmount-1) + ")"
                # mixSheet.cell(n, 13).value = round(poolDict[rPool]["weight"], 4)
                mixSheet.cell(n, 13).value = "=sum(J" + str(poolNumStartIndex) + ":J" + str(poolNumStartIndex+poolSampleAmount-1) + ")"
                # mixSheet.cell(n, 14).value = round(poolDict[rPool]["Fc"], 4)
                mixSheet.cell(n, 14).value = "=M" + str(poolNumStartIndex) + "/L" + str(poolNumStartIndex)
                mixSheet.cell(n, 14).number_format = '0.000'
                mixSheet.cell(n, 15).value = rPool

            lastPool = rPool
            poolNumStartIndexOrd += 1

            mixSheet.cell(n, 11).border = thin_border
            mixSheet.cell(n, 12).border = thin_border
            mixSheet.cell(n, 13).border = thin_border
            mixSheet.cell(n, 14).border = thin_border
            mixSheet.cell(n, 15).border = thin_border
        mixSheet.cell(n, 8).value = "=$" + chr(poolNumStartIndexOrd) + "$2/F" + str(n) + "*E" + str(n)
        mixSheet.cell(n, 8).number_format = '0.0'
        mixSheet.cell(n, 8).fill = fills[0]
        mixSheet.cell(n, 8).font = fonts[0]
        n += 1

    # 写入校正参数
    # print(poolList)
    # print(poolingFixNum)
    for pfn in range(len(poolingFixNum)):
        mixSheet.cell(1, 12+pfn).value = poolList[pfn]
        mixSheet.cell(2, 12+pfn).value = poolingFixNum[pfn]

    ################## 混样表 ####################################
    hunYang = poolingTemplate["混样"]
    hunYang.cell(3, 1).value = getTime()
    n = 6
    # hunYang.cell(n, 11).value = round(poolDataAmountAll, 2)
    # hunYang.cell(n, 12).value = round(poolVolAll, 1)
    # hunYang.cell(n, 13).value = round(poolWeightAll, 4)
    # hunYang.cell(n, 14).value = round(theoreticalPoolFc, 4)

    hunYang.cell(n, 11).border = thin_border
    hunYang.cell(n, 12).border = thin_border
    hunYang.cell(n, 13).border = thin_border
    hunYang.cell(n, 14).border = thin_border

    hunYang.cell(2, 12).value = fixNum

    for p in poolDict:
        hunYang.cell(n, 1).value = p
        hunYang.cell(n, 2).value = p
        hunYang.cell(n, 3).value = poolDict[p]["Fc"]
        hunYang.cell(n, 4).value = librarySize
        hunYang.cell(n, 5).value = poolDict[p]["dataAmount"]
        # hunYang.cell(n, 6).value = poolDict[p]["molFc"]
        hunYang.cell(n, 6).value = "=C" + str(n) + "*1515/D" + str(n)
        hunYang.cell(n, 7).value = p
        # hunYang.cell(n, 8).value = round(poolDict[p]["pooling"], 1)
        hunYang.cell(n, 8).value = "=$L$2/F" + str(n) + "*E" + str(n)
        # hunYang.cell(n, 10).value = round(poolDict[p]["poolWeight"], 4)
        hunYang.cell(n, 10).value = "=C" + str(n) + "*H" + str(n)

        hunYang.cell(n, 1).border = thin_border
        hunYang.cell(n, 2).border = thin_border
        hunYang.cell(n, 3).border = thin_border
        hunYang.cell(n, 4).border = thin_border
        hunYang.cell(n, 5).border = thin_border
        hunYang.cell(n, 6).border = thin_border
        hunYang.cell(n, 7).border = thin_border
        hunYang.cell(n, 8).border = thin_border
        hunYang.cell(n, 9).border = thin_border
        hunYang.cell(n, 10).border = thin_border
        
        n += 1

    hunYang.cell(6, 11).value = "=sum(E6:E" + str(n) + ")"
    hunYang.cell(6, 12).value = "=sum(H6:H" + str(n) + ")"
    hunYang.cell(6, 13).value = "=sum(J6:J" + str(n) + ")"
    hunYang.cell(6, 14).value = "=M6/L6"

    poolingTemplate.save(outputExcel)
    print("【完成】", getTime())


# main("测试3_实验信息.xlsx", "测试4_实验信息.xlsx", "中班", "晚班", "test.xlsx")

# GUI
sg.ChangeLookAndFeel("GreenTan")
report_layout = [
    [sg.Text("tNGS Pooler                                                                                         Version 0.7 20221119")],
    [sg.Text("实验表1"), sg.Input(key="classXlsx1", size=(70, 6)), sg.FileBrowse()],
    [sg.Text("实验表2"), sg.Input(key="classXlsx2", size=(70, 6)), sg.FileBrowse()],
    [sg.Text("表1描述"), sg.Input(key="class1Sig", size=(33, 6), default_text="中班"), sg.Text("表2描述"), sg.Input(key="class2Sig", size=(33, 6), default_text="晚班")],
    # [sg.Text("表1文库浓度列"), sg.Input(key="libCol1", size=(30, 6), default_text="M"), sg.Text("表2文库浓度列"), sg.Input(key="libCol2", size=(30, 6), default_text="M")],
    [sg.Text("输出表格名称"), sg.Input(key="outputXlsx", size=(70, 6)), sg.FileBrowse()],
    [sg.Text("运行信息"), sg.Output(key="log", size=(80, 10))],
    [sg.Button("运行"), sg.Button("退出")]
]

window = sg.Window("tNGS Pooler", report_layout)
while True:
    event, runningDict = window.read()
    if event == sg.WIN_CLOSED or event == "退出":
        break
    else:
        c1 = runningDict["classXlsx1"]
        if c1 == None:
            c1 = ""
        c2 = runningDict["classXlsx2"]
        if c2 == None:
            c2 = ""
        s1 = runningDict["class1Sig"]
        s2 = runningDict["class2Sig"]
        o = runningDict["outputXlsx"]
        try:
            main(c1, c2, s1, s2, o)
        except:
            print("【错误】请确认输入内容存在")
window.close()

