# -*- encoding: utf-8 -*-
# pzw
# 20210604
# 药物基因组学

import os
import sys
import argparse
import openpyxl
import xlrd
import time
from WordWriter3 import WordWriter
import PySimpleGUI as sg
import warnings

################## 程序主体 ######################

# 获得程序运行路径
def getAbsPath():
    now = os.path.abspath(os.path.dirname(sys.argv[0]))
    return now

# xls读取为xlsx
def xls2xlsx(xlsFile):
    book = xlrd.open_workbook(xlsFile)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1
    book_xlsx = openpyxl.workbook.Workbook()
    sheet_xlsx = book_xlsx.active
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet_xlsx.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)
    return book_xlsx

def main(sampleInfo, excelInput, outputDir):
    # 获得数据库路径
    databases = getAbsPath() + "/data"
    temp_asp_clo = databases + "/Aspirin_Clopidogrel.docx"
    temp_clo = databases + "/Clopidogrel.docx"
    temp_war = databases + "/Warfarin.docx"

    # 输出文件夹自动生成
    if not os.path.exists(outputDir):
        os.makedirs(outputDir)

    # 样本信息导入
    openpyxl.Workbook.encoding = "utf-8"
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        if sampleInfo.split(".")[-1] == "xlsx":
            sample_infos = openpyxl.load_workbook(sampleInfo)
        elif sampleInfo.split(".")[-1] == "xls":
            sample_infos = xls2xlsx(sampleInfo)
        else:
            print("未知样本信息文件拓展名")

    sample_infos = sample_infos
    samplesheet = sample_infos.active
    sample_dict = {}
    n = 2
    while n >= 2:
        # 基本信息
        sampleID = samplesheet["A"+str(n)].value
        if sampleID == None:
            break
        else:
            sampleID = samplesheet["A"+str(n)].value
            name = samplesheet["B"+str(n)].value if samplesheet["B"+str(n)].value != None else ""
            collect_date = samplesheet["C"+str(n)].value if samplesheet["C"+str(n)].value != None else ""
            if collect_date != "":
                y = collect_date.split("-")[0] + "年"
                m = collect_date.split("-")[1]
                m = m + "月"
                d = collect_date.split("-")[2]
                d = d + "日"
                collect_date = y + m + d
            send = samplesheet["E"+str(n)].value if samplesheet["E"+str(n)].value != None else ""
            doctor = samplesheet["J"+str(n)].value if samplesheet["J"+str(n)].value != None else ""
            office = samplesheet["L"+str(n)].value if samplesheet["L"+str(n)].value != None else ""
            age = samplesheet["M"+str(n)].value if samplesheet["M"+str(n)].value != None else ""
            gender = samplesheet["O"+str(n)].value if samplesheet["O"+str(n)].value != None else ""
            diagnosis = samplesheet["T"+str(n)].value if samplesheet["T"+str(n)].value != None else ""
            sample_dict[sampleID] = [name, collect_date, send, doctor, office, age, gender, diagnosis]
            n += 1

    # 遍历excel信息
    book = openpyxl.load_workbook(excelInput)
    sheet = book.active
    n = 3
    while n >= 3:
        project = sheet["A"+str(n)].value
        if project == None:
            break
        else:
            # 报告结果汇总
            sampleID = sheet["C"+str(n)].value
            if sampleID == None:
                print("样本条码未填写")
                break
            elif not str(sampleID) in sample_dict.keys():
                print("样本信息中未能找到对应样本编号")
                break
            else:
                resultsMap = {}
                resultsMap["#[name]#"] = sample_dict[sampleID][0]
                resultsMap["#[send]#"] = sample_dict[sampleID][2]
                resultsMap["#[collect_date]#"] = sample_dict[sampleID][1]
                resultsMap["#[report_date]#"] = time.strftime("%Y年") + time.strftime("%m月") + time.strftime("%d日")
                resultsMap["#[report_date2]#"] = time.strftime("%Y-%m-%d")
                resultsMap["#[report_date3]#"] = time.strftime("%Y.%m.%d")
                resultsMap["#[gender]#"] = sample_dict[sampleID][6]
                resultsMap["#[age]#"] = sample_dict[sampleID][5]
                resultsMap["#[types]#"] = sheet["D"+str(n)].value if sheet["D"+str(n)].value != None else ""
                resultsMap["#[diagnosis]#"] = sample_dict[sampleID][7]
                resultsMap["#[office]#"] = sample_dict[sampleID][4]
                resultsMap["#[doctor]#"] = sample_dict[sampleID][3]
                resultsMap["#[sampleID]#"] = str(sampleID)
                resultsMap["#[tester]#"] = sheet["E"+str(n)].value if sheet["E"+str(n)].value != None else ""
                resultsMap["#[TBS-name]#"] = resultsMap["#[name]#"]
                resultsMap["#[TBS-send]#"] = resultsMap["#[send]#"]
                resultsMap["#[TBS-collect_date]#"] = resultsMap["#[collect_date]#"]
                resultsMap["#[TBS-gender]#"] = resultsMap["#[gender]#"]
                resultsMap["#[TBS-age]#"] = resultsMap["#[age]#"]
                resultsMap["#[TBS-types]#"] = resultsMap["#[types]#"]
                resultsMap["#[TBS-diagnosis]#"] = resultsMap["#[diagnosis]#"]
                resultsMap["#[TBS-office]#"] = resultsMap["#[office]#"]
                resultsMap["#[TBS-doctor]#"] = resultsMap["#[doctor]#"]
                resultsMap["#[TBS-sampleID]#"] = resultsMap["#[sampleID]#"]
                resultsMap["#[Tester]#"] = resultsMap["#[tester]#"]
            
            # 分项目获取信息
            ## 阿司匹林,氯毗格雷
            if project == "阿司匹林,氯毗格雷":
                print("正在分析 " + str(sampleID) + "-" + sample_dict[sampleID][0] + " 项目：" + project)
                GPIIIa = sheet["F"+str(n)].value
                PEAR1 = sheet["G"+str(n)].value
                PTGS1 = sheet["H"+str(n)].value
                GSTP1 = sheet["I"+str(n)].value
                CYP2C19_2 = sheet["J"+str(n)].value
                CYP2C19_3 = sheet["K"+str(n)].value
                CYP2C19_17 = sheet["L"+str(n)].value

                # 报告分析
                ## 阿司匹林部分
                asp_data = open(databases + "/Aspirin.txt", "r", encoding="utf-8")
                asp_dict = {}
                for line in asp_data:
                    if not line.startswith("基因"):
                        lines = line.replace("\n", "").split("\t")
                        asp_dict[lines[0] + "-" + lines[1]] = [lines[2], lines[3]]
                asp_data.close()

                ### 抵抗风险
                res_risk = asp_dict["GPⅢa-" + GPIIIa][0] + asp_dict["PEAR1-" + PEAR1][0] + asp_dict["PTGS1-" + PTGS1][0]
                if "重度抵抗" in res_risk:
                    res_risk = "重度抵抗风险"
                elif "中度抵抗" in res_risk:
                    res_risk = "中度抵抗风险"
                else:
                    res_risk = "较低抵抗风险"

                ### 出血风险
                if GSTP1 == "AA":
                    asp_blood_risk = "较低出血风险"
                else:
                    asp_blood_risk = asp_dict["GSTP1-" + GSTP1][1]

                ### 结论
                asp_conclusion1 = res_risk + "，" + asp_blood_risk
                if not "中" in asp_conclusion1:
                    if not "重" in asp_conclusion1:
                        asp_conclusion1 = "较低风险"
                if PEAR1 == "GA" and PTGS1 == "AA" and GSTP1 == "AA":
                    asp_conclusion2 = "对阿司匹林应答不佳，需增加剂量。具体请结合临床实际选择治疗方案"
                else:
                    asp_conclusion2 = "请结合临床实际选择治疗方案"

                ## 氯毗格雷部分
                clo_data = open(databases + "/Clopidogrel.txt", "r", encoding="utf-8")
                clo_dict = {}
                for line in clo_data:
                    if not line.startswith("基因"):
                        lines = line.replace("\n", "").split("\t")
                        clo_dict[lines[3]] = [lines[0], lines[1], lines[2]]
                clo_data.close()
                clo_results = CYP2C19_2 + "/" + CYP2C19_3 + "/" + CYP2C19_17

                # 检测结果
                resultsMap["#[TBS-GPIIIa]#"] = GPIIIa
                resultsMap["#[TBS-PEAR1]#"] = PEAR1
                resultsMap["#[TBS-PTGS1]#"] = PTGS1
                resultsMap["#[TBS-GSTP1]#"] = GSTP1
                resultsMap["#[TBS-GPIIIa_type]#"] = "野生型" if GPIIIa == "PLA1/A1" else "突变型"
                resultsMap["#[TBS-PEAR1_type]#"] = "野生型" if PEAR1 == "GG" else "突变型"
                resultsMap["#[TBS-PTGS1_type]#"] = "野生型" if PTGS1 == "AA" else "突变型"
                resultsMap["#[TBS-GSTP1_type]#"] = "野生型" if GSTP1 == "AA" else "突变型"
                resultsMap["#[TBS-GPIIIa_risk]#"] = asp_dict["GPⅢa-" + GPIIIa][0]
                resultsMap["#[TBS-PEAR1_risk]#"] = asp_dict["PEAR1-" + PEAR1][0]
                resultsMap["#[TBS-PTGS1_risk]#"] = asp_dict["PTGS1-" + PTGS1][0]
                resultsMap["#[TBS-GSTP1_risk]#"] = asp_dict["GSTP1-" + GSTP1][0]
                resultsMap["#[TBS-GPIIIa_sig]#"] = asp_dict["GPⅢa-" + GPIIIa][1]
                resultsMap["#[TBS-PEAR1_sig]#"] = asp_dict["PEAR1-" + PEAR1][1]
                resultsMap["#[TBS-PTGS1_sig]#"] = asp_dict["PTGS1-" + PTGS1][1]
                resultsMap["#[TBS-GSTP1_sig]#"] = asp_dict["GSTP1-" + GSTP1][1]
                resultsMap["#[AspirinTypes]#"] = "GPⅢa({})，PEAR1({})，PTGS1({})，GSTP1({})".format(GPIIIa, PEAR1, PTGS1, GSTP1)
                resultsMap["#[asp_conclusion1]#"] = asp_conclusion1
                resultsMap["#[asp_conclusion2]#"] = asp_conclusion2
                resultsMap["#[TBS-CYP2C19_2]#"] = CYP2C19_2
                resultsMap["#[TBS-CYP2C19_3]#"] = CYP2C19_3
                resultsMap["#[TBS-CYP2C19_17]#"] = CYP2C19_17
                resultsMap["#[TBS-Clo]#"] = clo_dict[clo_results][2]
                resultsMap["#[Clo_type]#"] = clo_dict[clo_results][0]

                # 输出报告
                WordWriter(temp_asp_clo, outputDir + "/" + str(sampleID) + "-" + sample_dict[sampleID][0] + "-" + project + ".docx", resultsMap)
                # try:
                #     os.system("pandoc {} -f docx -o {}.pdf".format(outputDir + "/" + str(sampleID) + "-" + name + ".docx", outputDir + "/" + str(sampleID) + "-" + name))
                # except:
                #     print("请安装pandoc和win32TeX进行自动化pdf输出：")
                #     print(" https://pandoc.org/installing.html")
                #     print(" http://w32tex.org/index-zh.html")
                print(sampleID, "分析完成")

            ## 氯毗格雷
            elif project == "氯毗格雷":
                print("正在分析 " + str(sampleID) + "-" + sample_dict[sampleID][0] + " 项目：" + project)
                CYP2C19_2 = sheet["J"+str(n)].value
                CYP2C19_3 = sheet["K"+str(n)].value
                CYP2C19_17 = sheet["L"+str(n)].value
                
                clo_data = open(databases + "/Clopidogrel.txt", "r", encoding="utf-8")
                clo_dict = {}
                for line in clo_data:
                    lines = line.replace("\n", "").split("\t")
                    clo_dict[lines[3]] = [lines[0], lines[1], lines[2]]
                clo_data.close()
                clo_results = CYP2C19_2 + "/" + CYP2C19_3 + "/" + CYP2C19_17            
                resultsMap["#[TBS-CYP2C19_2]#"] = CYP2C19_2
                resultsMap["#[TBS-CYP2C19_3]#"] = CYP2C19_3
                resultsMap["#[TBS-CYP2C19_17]#"] = CYP2C19_17
                resultsMap["#[TBS-Clo]#"] = clo_dict[clo_results][2]
                resultsMap["#[Clo_type]#"] = clo_dict[clo_results][0]

                # 输出报告
                WordWriter(temp_clo, outputDir + "/" + str(sampleID) + "-" + sample_dict[sampleID][0] + "-" + project + ".docx", resultsMap)
                print(sampleID, "分析完成")

            ## 华法林
            elif project == "华法林":
                print("正在分析 " + str(sampleID) + "-" + sample_dict[sampleID][0] + " 项目：" + project)
                CYP2C9_3 = sheet["M"+str(n)].value
                VKORC1 = sheet["N"+str(n)].value
                if CYP2C9_3 == "AA":
                    CYP2C9_types = "*1/*1"
                elif CYP2C9_3 == "AC":
                    CYP2C9_types = "*1/*3"
                elif CYP2C9_3 == "CC":
                    CYP2C9_types = "*3/*3"
                else:
                    CYP2C9_types = ""
                    print("CYP2C9*3填写错误")
                war_data = open(databases + "/Warfarin.txt", "r", encoding="utf-8")
                war_dict = {}
                for line in war_data:
                    if not line.startswith("CYP"):
                        lines = line.replace("\n", "").split("\t")
                        war_dict[lines[0] + lines[1]] = lines[2]
                war_data.close()
                war_sig = war_dict[CYP2C9_3 + VKORC1]
                resultsMap["#[TBS-CYP2C9_3]#"] = CYP2C9_3
                resultsMap["#[TBS-CYP2C9_3_type]#"] = CYP2C9_types
                resultsMap["#[TBS-VKORC1]#"] = VKORC1
                resultsMap["#[TBS-War_sig]#"] = war_sig

                # 输出报告
                WordWriter(temp_war, outputDir + "/" + str(sampleID) + "-" + sample_dict[sampleID][0] + "-" + project + ".docx", resultsMap)
                print(sampleID, "分析完成")

            else:
                print("未找到该项目，请确认项目名称：", project)
            n += 1

# 命令行模式
# if __name__ == "__main__":
#     parser = argparse.ArgumentParser(
#         description="Report CVDs Drug @pzweuj\n\
#         qPCR 心脑血管用药报告\n\
#         目前已完成： \n\
#             阿司匹林,氯毗格雷\n\
#             氯毗格雷\n\
#             华法林",
#         prog="Report_CVDs_Drug.py",
#         usage="python Report_CVDs_Drug.py [-h] -s <sampleInfo> -i <excelInput> -o <outputDir>",
#         formatter_class=argparse.RawTextHelpFormatter
#     )
#     parser.add_argument("-v", "--version", action="version",
#         version="Version 0.4 20210604")
#     parser.add_argument("-s", "--sample", type=str,
#         help="样本信息文件")
#     parser.add_argument("-i", "--rawdata", type=str,
#         help="excel信息文件")
#     parser.add_argument("-o", "--output", type=str,
#         help="输出文件夹")

#     if len(sys.argv[1:]) == 0:
#         parser.print_help()
#         parser.exit()
#     args = parser.parse_args()
#     main(sampleInfo=args.sample, excelInput=args.rawdata, outputDir=args.output)


###################  GUI ##########################
sg.theme("Dark Blue 3")
layout = [
    [sg.Text("药物基因组学报告生成                                                                                                    Version 0.4 20210604")],
    [sg.Text("样本信息文件"), sg.Input(key="SampleInfoFile", size=(70, 6)), sg.FileBrowse()],
    [sg.Text("检测结果文件"), sg.Input(key="ExcelInputFile", size=(70, 6)), sg.FileBrowse()],
    [sg.Text(" 输出文件夹 "), sg.Input(key="OutputDirPath", size=(70, 6)), sg.FolderBrowse()],
    [sg.Text("运行信息"), sg.Output(key="log", size=(80, 10))],
    [sg.Button("运行"), sg.Button("退出")]
]

window = sg.Window("qPCR 心脑血管报告", layout)
while True:
    event, runningDict = window.read()
    if event == sg.WIN_CLOSED or event == "退出":
        break
    else:
        s = runningDict["SampleInfoFile"]
        e = runningDict["ExcelInputFile"]
        o = runningDict["OutputDirPath"]
        try:
            main(s, e, o)
        except:
            print("请确认输入内容存在")

window.close()