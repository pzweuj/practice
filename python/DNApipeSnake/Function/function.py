# coding=utf-8
# pzweuj
# 20210511
# DNApipeLine Snakemake Function

import os
import sys
import json
from openpyxl import Workbook

# 创建文件夹
def mkdir(folder_path):
    folder = os.path.exists(folder_path)
    if not folder:
        os.makedirs(folder_path)
        print("Create new folder " + folder_path + " done!")
    else:
        print("Folder " + folder_path + " already exists!")


# 注释结果整理
def refTranscript(refTranscriptFile):
    refT = open(refTranscriptFile, "r")
    refDict = {}
    for r in refT:
        if not r.startswith("#"):
            gene = r.split("\t")[0]
            ts = r.split("\t")[1]
            refDict[gene] = ts
    refT.close()
    return refDict

def ResultsFilter(annvarResultsFile, annvarFixFile, refTranscriptFile):
    buildver = "hg19"

    annvarResults = open(annvarResultsFile, "r")
    results = open(annvarFixFile, "w")
    results.write("Chr\tStart\tEnd\tRef\tAlt\tGene\tType\tTranscript\tcHGVS\tpHGVS\tVAF\tConsequence\t"\
        "AffectedExon\tDepth\tAlt_AD\tFunc.refGene\tAAChange.refGene\tavsnp150\tAF\tAF_popmax\tAF_male\t"\
        "AF_female\tAF_eas\tCLNDN\tCLNSIG\tInterVar_automated\tInterVar_sig\tJax_Ckb_Variants_Summary\tJax_Ckb_Drug_Summary\t"\
        "Civic\tOncoKB\tcosmic92_coding\tDamagePredCount\tSIFT_pred\tPolyphen2_HDIV_pred\tPolyphen2_HVAR_pred\t"\
        "LRT_pred\tMutationTaster_pred\tMutationAssessor_pred\tFATHMM_pred\tPROVEAN_pred\t"\
        "M-CAP_pred\tREVEL_score\n")

    # 生成字典keys
    lineDict = {}
    for line in annvarResults:
        if line.startswith("Chr\tStart"):
            lines = line.replace("\n", "").split("\t")
            for k in lines:
                lineDict[k] = "-"
    annvarResults.close()

    # 再次打开
    annvarResults = open(annvarResultsFile, "r")
    for line in annvarResults:
        if not line.startswith("Chr\tStart"):
            lines = line.replace("\n", "").split("\t")
            k = list(lineDict.keys())
            
            # 生成字典
            for i in range(len(lines)):
                lineDict[k[i]] = lines[i]
            
            # 开始处理
            # 处理Insertion状态下的End
            if lineDict["Ref"] == "-":
                lineDict["Type"] = "Insertion"
                lineDict["End"] = "-"
            elif lineDict["Alt"] == "-":
                lineDict["Type"] = "Deletion"
            elif len(lineDict["Ref"]) == 1 and len(lineDict["Alt"]) == 1:
                lineDict["Type"] = "SNV"
            else:
                lineDict["Type"] = "Complex"

            # 处理基因，使用snpeff的注释信息
            otif8 = lineDict["Otherinfo8"].split(";")
            snpeff_anno = "ANN=?|N|N|N|N|transcript|N|N|N|N|N|N|N|N|N|"
            for o8 in otif8:
                if o8.startswith("ANN="):
                    snpeff_anno = o8
            snpeff_annos = snpeff_anno.replace("ANN=", "").split(",")
            
            # 此处需使用固定转录本，未完成
            if len(snpeff_annos) > 1:
                snpeff_select = snpeff_annos[0]
                gene = snpeff_select.split("|")[3]
                refTrans = refTranscript(refTranscriptFile)
                try:
                    ts = refTrans[gene]
                    for sa in snpeff_annos:
                        if ts.split(".")[0] in snpeff_annos:
                            if sa.split("|")[6] != "":
                                snpeff_select = sa
                except:
                    snpeff_select = snpeff_select
            else:
                snpeff_select = snpeff_annos[0]
            s = snpeff_select.split("|")
            for ss in range(len(s)):
                if s[ss] == "":
                    s[ss] = "-"

            lineDict["Gene"] = s[3]
            lineDict["Transcript"] = s[6]
            lineDict["cHGVS"] = s[9]
            lineDict["pHGVS"] = s[10]
            lineDict["AffectedExon"] = s[8]

            # 处理Consequence
            # print(s[1])
            if lineDict["Func.refGene"] == "exonic" and lineDict["Type"] == "Complex":
                if "_" in lineDict["pHGVS"]:
                    lineDict["Consequence"] = "Complex_mutation"
                elif "stop_gained" in s[1]:
                    lineDict["Consequence"] = "Nonsense_substitution"
                elif "missense" in s[1]:
                    lineDict["Consequence"] = "Missense_substitution"
                elif "synonymous" in s[1]:
                    lineDict["Consequence"] = "Synonymous_substitution"
                else:
                    lineDict["Consequence"] = "Complex_mutation"
            elif "missense" in s[1]:
                lineDict["Consequence"] = "Missense_substitution"
            elif "splice" in s[1]:
                lineDict["Consequence"] = "Splice_Site_mutation"
            elif "synonymous" in s[1]:
                lineDict["Consequence"] = "Synonymous_substitution"
            elif "inframe_deletion" in s[1]:
                lineDict["Consequence"] = "Inframe_deletion"
            elif "inframe_insertion" in s[1]:
                lineDict["Consequence"] = "Inframe_insertion"
            elif "frameshift" in s[1]:
                if lineDict["Type"] == "Insertion":
                    lineDict["Consequence"] = "Frameshift_insertion"
                elif lineDict["Type"] == "Deletion":
                    lineDict["Consequence"] = "Frameshift_deletion"
                else:
                    lineDict["Consequence"] = "Frameshift_variant"
            elif "stop_gained" in s[1]:
                lineDict["Consequence"] = "Nonsense_substitution"
            else:
                lineDict["Consequence"] = "Other"

            # 处理VAF
            otif10 = lineDict["Otherinfo10"].split(":")
            otif9 = lineDict["Otherinfo9"].split(":")        
            format_zip = list(zip(otif9, otif10))
            format_dict = {}
            for f in format_zip:
                format_dict[f[0]] = f[1]

            lineDict["DP"] = format_dict["DP"]
            lineDict["Alt_AD"] = format_dict["AD"]
            if not "," in lineDict["Alt_AD"]:
                lineDict["Alt_AD"] = lineDict["Alt_AD"]
            else:
                lineDict["Alt_AD"] = lineDict["Alt_AD"].split(",")[1]
            lineDict["Ref_AD"] = str(int(lineDict["DP"]) - int(lineDict["Alt_AD"]))

            try:
                AF = "%.2f" % ((float(lineDict["Alt_AD"]) / float(lineDict["DP"])) * 100) + "%"
            except Exception:
                AF = "-"
            lineDict["VAF"] = AF

            # 合并证据等级
            lineDict["InterVar_sig"] = "PVS1=" + lineDict["PVS1"] + ";" + \
                "PS1=" + lineDict["PS1"] + ";" + \
                "PS2=" + lineDict["PS2"] + ";" + \
                "PS3=" + lineDict["PS3"] + ";" + \
                "PS4=" + lineDict["PS4"] + ";" + \
                "PM1=" + lineDict["PM1"] + ";" + \
                "PM2=" + lineDict["PM2"] + ";" + \
                "PM3=" + lineDict["PM3"] + ";" + \
                "PM4=" + lineDict["PM4"] + ";" + \
                "PM5=" + lineDict["PM5"] + ";" + \
                "PM6=" + lineDict["PM6"] + ";" + \
                "PP1=" + lineDict["PP1"] + ";" + \
                "PP2=" + lineDict["PP2"] + ";" + \
                "PP3=" + lineDict["PP3"] + ";" + \
                "PP4=" + lineDict["PP4"] + ";" + \
                "PP5=" + lineDict["PP5"] + ";" + \
                "BA1=" + lineDict["BA1"] + ";" + \
                "BS1=" + lineDict["BS1"] + ";" + \
                "BS2=" + lineDict["BS2"] + ";" + \
                "BS3=" + lineDict["BS3"] + ";" + \
                "BS4=" + lineDict["BS4"] + ";" + \
                "BP1=" + lineDict["BP1"] + ";" + \
                "BP2=" + lineDict["BP2"] + ";" + \
                "BP3=" + lineDict["BP3"] + ";" + \
                "BP4=" + lineDict["BP4"] + ";" + \
                "BP5=" + lineDict["BP5"] + ";" + \
                "BP6=" + lineDict["BP6"] + ";" + \
                "BP7=" + lineDict["BP7"]

            # 整理输出结果
            output = [lineDict["Chr"], lineDict["Start"], lineDict["End"], lineDict["Ref"],
                lineDict["Alt"], lineDict["Gene"], lineDict["Type"], lineDict["Transcript"],
                lineDict["cHGVS"], lineDict["pHGVS"], lineDict["VAF"], lineDict["Consequence"],
                lineDict["AffectedExon"], lineDict["DP"], lineDict["Alt_AD"], lineDict["Func.refGene"],
                lineDict["AAChange.refGene"], lineDict["avsnp150"], lineDict["AF"], lineDict["AF_popmax"], lineDict["AF_male"], 
                lineDict["AF_female"], lineDict["AF_eas"], lineDict["CLNDN"], lineDict["CLNSIG"],
                lineDict["InterVar_automated"], lineDict["InterVar_sig"], lineDict["Jax_Ckb_Variants_Summary"],
                lineDict["Jax_Ckb_Drug_Summary"], lineDict["Civic"], lineDict["OncoKB"], lineDict["cosmic92_coding"],
                lineDict["DamagePredCount"], lineDict["SIFT_pred"], lineDict["Polyphen2_HDIV_pred"],
                lineDict["Polyphen2_HVAR_pred"], lineDict["LRT_pred"], lineDict["MutationTaster_pred"],
                lineDict["MutationAssessor_pred"], lineDict["FATHMM_pred"], lineDict["PROVEAN_pred"],
                lineDict["M-CAP_pred"], lineDict["REVEL_score"]
            ]

            print("\t".join(output))
            results.write("\t".join(output) + "\n")
    print("完成annovar结果过滤与格式调整")

# CNV结果过滤与整理
def cnvkit_filter(CallCnsFile, FilterResults):
    # 需要建立hg19坐标与外显子编号对应数据库，落实是哪个外显子的扩增

    cnvkitResultsFile = open(CallCnsFile, "r")
    cnvkitResults = open(FilterResults, "w")
    cnvkitResults.write("chromsome\tstart\tend\tgene\tVAF\tlog2\tdepth\tp_ttest\tprobes\tweight\n")
    for line in cnvkitResultsFile:
        if line.startswith("chromosome"):
            continue
        else:
            lineAfterSplit = line.split("\n")[0].split("\t")
            chrom = lineAfterSplit[0]
            start = lineAfterSplit[1]
            end = lineAfterSplit[2]
            gene = lineAfterSplit[3]
            log2 = lineAfterSplit[4]
            VAF = lineAfterSplit[5]
            depth = lineAfterSplit[6]
            if len(lineAfterSplit) == 10:
                p_ttest = lineAfterSplit[7]
                probes = lineAfterSplit[8]
                weight = lineAfterSplit[9]
            else:
                p_ttest = "-"
                probes = lineAfterSplit[7]
                weight = lineAfterSplit[8]

            if gene == "-":
                continue
            elif VAF == "2":
                continue
            else:
                filter_output = [chrom, start, end, gene, VAF, log2, depth, p_ttest, probes, weight]
                outputString = "\t".join(filter_output)
                cnvkitResults.write(outputString + "\n")
    cnvkitResults.close()
    cnvkitResultsFile.close()

# Fusion结果过滤与注释
def lumpy_filter(lumpyResultsFile, lumpyFilterFile, bamFile, filterDP, filterVAF, samtools):
    filterAlt = int(filterDP * filterVAF)
    lumpyResults = open(lumpyResultsFile, "r")
    lumpyResultsFilter = open(lumpyFilterFile, "w")
    for line in lumpyResults:
        if line.startswith("#"):
            lumpyResultsFilter.write(line)
        else:
            if not "SVTYPE=BND" in line:
                continue

            lineAfterSplit = line.split("\n")[0].split("\t")
            chrom1 = lineAfterSplit[0]
            pos1 = lineAfterSplit[1]
            alt = lineAfterSplit[4]
            if alt[0:2] == "N[":
                chrom2s = alt.replace("N[", "").replace("[", "")
            elif alt[0:2] == "N]":
                chrom2s = alt.replace("N]", "").replace("]", "")
            elif alt[-2:] == "[N":
                chrom2s = alt.replace("[N", "").replace("[", "")
            elif alt[-2:] == "]N":
                chrom2s = alt.replace("]N", "").replace("]", "")
            else:
                continue

            chrom2 = chrom2s.split(":")[0]
            pos2 = chrom2s.split(":")[1]

            FORMAT = "PR:SR"
            FORMAT_origin = lineAfterSplit[8].split(":")
            FORMAT_info = lineAfterSplit[9]
            infos = FORMAT_info.split(":")
            formatDict = {}
            for F in range(len(FORMAT_origin)):
                formatDict[FORMAT_origin[F]] = infos[F]
            try:
                PR_alt = formatDict["PE"]
            except:
                PR_alt = "0"
            try:
                SR_alt = formatDict["SR"]
            except:
                SR_alt = "0"
            check1 = chrom1 + ":" + pos1 + "-" + pos1
            check2 = chrom2 + ":" + pos2 + "-" + pos2
            
            # 突变reads数过滤
            if (int(PR_alt) + int(SR_alt)) < filterAlt:
                continue

            dp1 = os.popen("{samtools} depth {bamFile} -r {check1}".format(samtools=samtools, bamFile=bamFile, check1=check1))
            dp2 = os.popen("{samtools} depth {bamFile} -r {check2}".format(samtools=samtools, bamFile=bamFile, check2=check2))
            for d1 in dp1.readlines():
                if d1.startswith("chr"):
                    depth1 = d1.replace("\n", "").split("\t")[2]
            for d2 in dp2.readlines():
                if d2.startswith("chr"):
                    depth2 = d2.replace("\n", "").split("\t")[2]
            DP = int(depth1) + int(depth2)

            # 深度过滤
            if DP < filterDP:
                continue

            print(line)
            PR_ref = str(DP)
            SR_ref = "0"

            output = "\t".join(lineAfterSplit[0:8]) + "\t" + FORMAT + "\t" + PR_ref + "," + PR_alt + ":" + SR_ref + "," + SR_alt + "\n"
            lumpyResultsFilter.write(output)
    lumpyResultsFilter.close()
    lumpyResults.close()

def manta_filter(mantaResultsFile, mantaFilter, filterDP, filterVAF):
    filterAlt = int(filterDP * filterVAF)

    mantaResults = open(mantaResultsFile, "r")
    mantaResultsFilter = open(mantaFilter, "w")
    for line in mantaResults:
        if line.startswith("#"):
            mantaResultsFilter.write(line)
        else:
            lineAfterSplit = line.split("\n")[0].split("\t")
            FORMAT = lineAfterSplit[8]
            
            FORMAT_info = lineAfterSplit[9]

            if "PR" not in FORMAT:
                SR = FORMAT_info.split(",")
                PR_ref = "0"
                PR_alt = "0"
                SR_ref = SR[0]
                SR_alt = SR[1]
            elif "SR" not in FORMAT:
                PR = FORMAT_info.split(",")
                SR_ref = "0"
                SR_alt = "0"
                PR_ref = PR[0]
                PR_alt = PR[1]
            else:
                FORMAT_infos = FORMAT_info.split(":")
                PR = FORMAT_infos[0].split(",")
                SR = FORMAT_infos[1].split(",")
                PR_ref = PR[0]
                PR_alt = PR[1]
                SR_ref = SR[0]
                SR_alt = SR[1]
            
            Ref_total = int(PR_ref) + int(SR_ref)
            Alt_total = int(PR_alt) + int(SR_alt)
            if Ref_total < filterDP:
                continue
            if Alt_total < filterAlt:
                continue

            mantaResultsFilter.write(line)
    mantaResultsFilter.close()
    mantaResults.close()


# 自编sv注释
# 对数据库进行解析，数据库使用refFlat，可通过ucsc直接下载
def checkFusionHotSpot(database, fusionStrand, chrom, breakPoint):

    db = open(database, "r")
    output = []
    ts_output = []
    for line in db:
        l = line.split("\n")[0].split("\t")
        gene = l[0]
        ts = l[1]
        chrom_db = l[2]
        strand = l[3]
        gene_start = l[4]
        gene_end = l[5]
        cds_start = l[6]
        cds_end = l[7]
        exon_nums = l[8]
        exon_starts = l[9]
        exon_ends = l[10]

        # 寻找注释
        if chrom == chrom_db:
            # 找到目标基因
            if (int(breakPoint) >= int(gene_start)) and (int(breakPoint) <= int(gene_end)):

                exon_starts_list = exon_starts.split(",")[0: -1]
                exon_ends_list = exon_ends.split(",")[0: -1]

                # 将外显子起止点形成列表
                exon_checkpoint = []
                i = 0
                while i < int(exon_nums):
                    exon_checkpoint.append(int(exon_starts_list[i]))
                    exon_checkpoint.append(int(exon_ends_list[i]))
                    i += 1

                # 判断目标基因的转录方向
                if strand == "+":
                    if int(breakPoint) < exon_checkpoint[0]:
                        if fusionStrand == "+":
                            exon_out = gene + "_5UTR"
                        elif fusionStrand == "-":
                            exon_out = gene + "_exon1"
                        else:
                            exon_out = gene + "_?"

                    elif int(breakPoint) > exon_checkpoint[-1]:
                        if fusionStrand == "+":
                            exon_out = gene + "_exon" + exon_nums
                        elif fusionStrand == "-":
                            exon_out = gene + "_3UTR"
                        else:
                            exon_out = gene + "_?"

                    else:
                        n = 0
                        while n < (2 * int(exon_nums) - 1):
                            if (int(breakPoint) >= exon_checkpoint[n]) and (int(breakPoint) <= exon_checkpoint[n+1]):
                                checkN = n
                            n += 1

                        if fusionStrand == "+":
                            exon_out = gene + "_exon" + str(checkN // 2 + 1)
                        elif fusionStrand == "-":
                            exon_out = gene + "_exon" + str((checkN + 1) // 2 + 1)
                        else:
                            exon_out = gene + "_?"


                if strand == "-":
                    if int(breakPoint) < exon_checkpoint[0]:
                        if fusionStrand == "+":
                            exon_out = gene + "_3UTR"
                        elif fusionStrand == "-":
                            exon_out = gene + "_exon" + exon_nums
                        else:
                            exon_out = gene + "_?"


                    elif int(breakPoint) > exon_checkpoint[-1]:
                        if fusionStrand == "+":
                            exon_out = gene + "_exon1"
                        elif fusionStrand == "-":
                            exon_out = gene + "_5UTR"
                        else:
                            exon_out = gene + "_?"

                    else:
                        n = 0
                        while n < (2 * int(exon_nums) - 1):
                            if (int(breakPoint) >= exon_checkpoint[n]) and (int(breakPoint) <= exon_checkpoint[n+1]):
                                checkN = n
                            n += 1

                        if fusionStrand == "+":
                            exon_out = gene + "_exon" + str(int(exon_nums) - checkN // 2)
                        elif fusionStrand == "-":
                            exon_out = gene + "_exon" + str(int(exon_nums) - (checkN + 1) // 2)
                        else:
                            exon_out = gene + "_?"

                output.append(exon_out)
                ts_output.append(ts)

    if not output:
        output.append("GeneUnknown_exon?")        
    if not ts_output:
        ts_output.append("Transcript?")
    db.close()
    return [output, ts_output]

def sv_anno(LumpyFilter, MantaFilter, fusionFile, refFlat):

    # lumpy
    svFile = open(LumpyFilter, "r")
    svAnno = open(fusionFile, "w")
    svAnno.write("chrom1\tbreakpoint1\tgene1\tchrom2\tbreakpoint2\tgene2\tfusionType\tAlt\tgeneSymbol\tPR\tSR\tDP\tVAF\tExon\tTranscript\tSource\n")
    for line in svFile:
        if line.startswith("#"):
            continue
        else:
            lineAfterSplit = line.split("\t")
            chrom = lineAfterSplit[0]
            Pos = lineAfterSplit[1]
            ID = lineAfterSplit[2]
            Ref = lineAfterSplit[3]
            Alt = lineAfterSplit[4]
            Qual = lineAfterSplit[5]
            Filter = lineAfterSplit[6]
            Info = lineAfterSplit[7]
            Format = lineAfterSplit[8]

            Sample = lineAfterSplit[9]

            if chrom == "chrM":
                continue

            if ":" in Format:
                S = Sample.split(":")
                PR = S[0].split(",")
                SR = S[1].split(",")
                PR_ref = int(PR[0])
                PR_alt = int(PR[1])
                SR_ref = int(SR[0])
                SR_alt = int(SR[1])
            else:
                PR = Sample.split(",")
                PR_ref = int(PR[0])
                PR_alt = int(PR[1])
                SR_ref = SR_alt = 0

            DP = PR_ref + SR_ref + PR_alt + SR_alt
            VAF = "%.2f" % (100 * float(PR_alt + SR_alt) / DP) + "%"

            if ("[" in Alt) or ("]" in Alt):
                if chrom in Alt:
                    fusionType = "Inversion"
                else:
                    fusionType = "Translocation"

                # G    [chr1:1111111[G
                if Alt[0] == "[":
                    chrom2 = Alt.split(":")[0].split("[")[1]
                    breakpoint2 = Alt.split(":")[1].split("[")[0]
                    strand = "--"
                
                # G    G[chr1:1111111[
                elif Alt[-1] == "[":
                    chrom2 = Alt.split(":")[0].split("[")[1]
                    breakpoint2 = Alt.split(":")[1].split("[")[0]
                    strand = "+-"

                # G    ]chr1:1111111]G
                elif Alt[0] == "]":
                    chrom2 = Alt.split(":")[0].split("]")[1]
                    breakpoint2 = Alt.split(":")[1].split("]")[0]
                    strand = "-+"

                # G    G]chr1:1111111]
                elif Alt[-1] == "]":
                    chrom2 = Alt.split(":")[0].split("]")[1]
                    breakpoint2 = Alt.split(":")[1].split("]")[0]
                    strand = "++"

                else:
                    chrom2 = "Unknown"
                    breakpoint2 = "Unknown"
                    strand = "Unknown"
                    print("can not analysis: " + ID)

                database = refFlat
                svStrand = strand[0]
                svChrom = chrom
                breakPoint = Pos
                anno1 = checkFusionHotSpot(database, svStrand, svChrom, breakPoint)
                if chrom2 != "Unknown":
                    svStrand = strand[1]
                    svChrom = chrom2
                    breakPoint = breakpoint2
                    anno2 = checkFusionHotSpot(database, svStrand, svChrom, breakPoint)
                else:
                    anno2 = [["GeneUnknown_exon?"], ["Transcript?"]]

                gene1 = anno1[0][0].split("_")[0]
                gene2 = anno2[0][0].split("_")[0]

                if gene1 == gene2:
                    continue

                exon_output = ",".join(anno1[0]) + "|" + ",".join(anno2[0])
                transcript_output = ",".join(anno1[1]) + "|" + ",".join(anno2[1])
                outputStringList = [chrom, Pos, gene1, chrom2, breakpoint2, gene2, fusionType, Alt, gene1 + "-" + gene2, str(PR_alt), str(SR_alt), str(DP), VAF, exon_output, transcript_output]

                outputString = "\t".join(outputStringList)
                print(outputString)
                svAnno.write(outputString + "\tLumpy\n")
    svFile.close()
    
    # manta
    svFile = open(MantaFilter, "r")
    for line in svFile:
        if line.startswith("#"):
            continue
        else:
            lineAfterSplit = line.replace("\n", "").split("\t")
            chrom = lineAfterSplit[0]
            Pos = lineAfterSplit[1]
            ID = lineAfterSplit[2]
            Ref = lineAfterSplit[3]
            Alt = lineAfterSplit[4]
            Qual = lineAfterSplit[5]
            Filter = lineAfterSplit[6]
            Info = lineAfterSplit[7]
            Format = lineAfterSplit[8]

            Sample = lineAfterSplit[9]

            if chrom == "chrM":
                continue

            if ":" in Format:
                S = Sample.split(":")
                PR = S[0].split(",")
                SR = S[1].split(",")
                PR_ref = int(PR[0])
                PR_alt = int(PR[1])
                SR_ref = int(SR[0])
                SR_alt = int(SR[1])
            else:
                PR = Sample.split(",")
                PR_ref = int(PR[0])
                PR_alt = int(PR[1])
                SR_ref = SR_alt = 0

            DP = PR_ref + SR_ref + PR_alt + SR_alt
            VAF = "%.2f" % (100 * float(PR_alt + SR_alt) / DP) + "%"

            if ("[" in Alt) or ("]" in Alt):
                if chrom in Alt:
                    fusionType = "Inversion"
                else:
                    fusionType = "Translocation"

                # G    [chr1:1111111[G
                if Alt[0] == "[":
                    chrom2 = Alt.split(":")[0].split("[")[1]
                    breakpoint2 = Alt.split(":")[1].split("[")[0]
                    strand = "--"
                
                # G    G[chr1:1111111[
                elif Alt[-1] == "[":
                    chrom2 = Alt.split(":")[0].split("[")[1]
                    breakpoint2 = Alt.split(":")[1].split("[")[0]
                    strand = "+-"

                # G    ]chr1:1111111]G
                elif Alt[0] == "]":
                    chrom2 = Alt.split(":")[0].split("]")[1]
                    breakpoint2 = Alt.split(":")[1].split("]")[0]
                    strand = "-+"

                # G    G]chr1:1111111]
                elif Alt[-1] == "]":
                    chrom2 = Alt.split(":")[0].split("]")[1]
                    breakpoint2 = Alt.split(":")[1].split("]")[0]
                    strand = "++"

                else:
                    chrom2 = "Unknown"
                    breakpoint2 = "Unknown"
                    strand = "Unknown"
                    print("can not analysis: " + ID)

                database = refFlat
                svStrand = strand[0]
                svChrom = chrom
                breakPoint = Pos

                anno1 = checkFusionHotSpot(database, svStrand, svChrom, breakPoint)
                if chrom2 != "Unknown":
                    svStrand = strand[1]
                    svChrom = chrom2
                    breakPoint = breakpoint2
                    anno2 = checkFusionHotSpot(database, svStrand, svChrom, breakPoint)
                else:
                    anno2 = [["GeneUnknown_exon?"], ["Transcript?"]]

                gene1 = anno1[0][0].split("_")[0]
                gene2 = anno2[0][0].split("_")[0]

                if gene1 == gene2:
                    continue

                exon_output = ",".join(anno1[0]) + "|" + ",".join(anno2[0])
                transcript_output = ",".join(anno1[1]) + "|" + ",".join(anno2[1])
                outputStringList = [chrom, Pos, gene1, chrom2, breakpoint2, gene2, fusionType, Alt, gene1 + "-" + gene2, str(PR_alt), str(SR_alt), str(DP), VAF, exon_output, transcript_output]

                outputString = "\t".join(outputStringList)
                print(outputString)
                svAnno.write(outputString + "\tManta\n")        
    svFile.close()        
    svAnno.close()


# Fastp结果整理
def fastp_filter(Sample, FastpJson, fastpReportFile):
    fastpJsonReport = FastpJson
    jsonFile = json.load(open(fastpJsonReport, "r"))

    rawReads = jsonFile["summary"]["before_filtering"]["total_reads"]
    rawBases = jsonFile["summary"]["before_filtering"]["total_bases"]
    rawQ20Bases = jsonFile["summary"]["before_filtering"]["q20_bases"]
    rawQ30Bases = jsonFile["summary"]["before_filtering"]["q30_bases"]
    rawQ20Rate = "%.2f" % (float(rawQ20Bases / rawBases) * 100) + "%"
    rawQ30Rate = "%.2f" % (float(rawQ30Bases / rawBases) * 100) + "%"
    rawGC = "%.2f" % (jsonFile["summary"]["before_filtering"]["gc_content"] * 100) + "%"
    duplicationRate = "%.2f" % (jsonFile["duplication"]["rate"] * 100) + "%"

    cleanReads = jsonFile["summary"]["after_filtering"]["total_reads"]
    cleanBases = jsonFile["summary"]["after_filtering"]["total_bases"]
    cleanQ20Bases = jsonFile["summary"]["after_filtering"]["q20_bases"]
    cleanQ30Bases = jsonFile["summary"]["after_filtering"]["q30_bases"]
    cleanQ20Rate = "%.2f" % (float(cleanQ20Bases / cleanBases) * 100) + "%"
    cleanQ30Rate = "%.2f" % (float(cleanQ30Bases / cleanBases) * 100) + "%"
    cleanGC = "%.2f" % (jsonFile["summary"]["after_filtering"]["gc_content"] * 100) + "%"

    fastpReport = open(fastpReportFile, "w")
    fastpReport.write(Sample + " fastp QC Report\n")
    fastpReport.write("rawReads\t" + str(rawReads) + "\n")
    fastpReport.write("rawBases\t" + str(rawBases) + "\n")
    fastpReport.write("raw_q20\t" + str(rawQ20Bases) + "\n")
    fastpReport.write("raw_q20_rate\t" + rawQ20Rate + "\n")
    fastpReport.write("raw_q30\t" + str(rawQ30Bases) + "\n")
    fastpReport.write("raw_q30_rate\t" + rawQ30Rate + "\n")
    fastpReport.write("raw_GC_content\t" + rawGC + "\n")
    fastpReport.write("duplicationRate\t" + duplicationRate + "\n")
    fastpReport.write("cleanReads\t" + str(cleanReads) + "\n")
    fastpReport.write("cleanBases\t" + str(cleanBases) + "\n")
    fastpReport.write("clean_q20\t" + str(cleanQ20Bases) + "\n")
    fastpReport.write("clean_q20_rate\t" + cleanQ20Rate + "\n")
    fastpReport.write("clean_q30\t" + str(cleanQ30Bases) + "\n")
    fastpReport.write("clean_q30_rate\t" + cleanQ30Rate + "\n")
    fastpReport.write("clean_GC_content\t" + cleanGC + "\n")
    fastpReport.close()


# 聚合结果
def read_txt(txt):
    txtFile = open(txt, "r", encoding="utf-8")
    l = []
    for i in txtFile:
        l.append(i.split("\n")[0])
    txtFile.close()
    return l

def write_excel(ws, contents):
    for i in range(len(contents)):
        col = contents[i].split("\t")
        for c in range(len(col)):
            cstring = col[c]
            cell = ws.cell(row=1+i, column=1+c)
            cell.value = cstring

# 结果合并到excel
def mergeResultsToExcel(resultsDir, sampleID):
    mkdir(resultsDir + "/results")
    dir_list = os.listdir(resultsDir)
    wb = Workbook()

    if "QC" in dir_list:
        qc = os.listdir(resultsDir + "/QC")
        if len(qc) != 0:
            for q in qc:
                if sampleID in q:
                    content = read_txt(resultsDir + "/QC/" + q)
                    ws_qc = wb.create_sheet(q.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_qc, content)

    if "annotation" in dir_list:
        anno = os.listdir(resultsDir + "/annotation")
        if len(anno) != 0:
            for a in anno:
                if sampleID in a:
                    if not "multianno" in a:
                        if "txt" in a:
                            content = read_txt(resultsDir + "/annotation/" + a)
                            ws_anno = wb.create_sheet(a.replace(".txt", "").replace(sampleID + ".", ""))
                            write_excel(ws_anno, content)

    if "cnv" in dir_list:
        cnv = os.listdir(resultsDir + "/cnv")
        if len(cnv) != 0:
            for c in cnv:
                if sampleID in c:
                    content = read_txt(resultsDir + "/cnv/" + c)
                    ws_cnv = wb.create_sheet(c.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_cnv, content)

    if "fusion" in dir_list:
        sv = os.listdir(resultsDir + "/fusion")
        if len(sv) != 0:
            for s in sv:
                if ".txt" in s:
                    if sampleID in s:
                        content = read_txt(resultsDir + "/fusion/" + s)
                        ws_sv = wb.create_sheet(s.replace(".txt", "").replace(sampleID + ".", ""))
                        write_excel(ws_sv, content)

    if "msi" in dir_list:
        msi = os.listdir(resultsDir + "/msi")
        if len(msi) != 0:
            for m in msi:
                if sampleID in m:
                    content = read_txt(resultsDir + "/msi/" + m)
                    ws_msi = wb.create_sheet(m.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_msi, content)

    if "HLA" in dir_list:
        hla = os.listdir(resultsDir + "/HLA")
        if len(hla) != 0:
            for h in hla:
                if sampleID in h:
                    content = read_txt(resultsDir + "/HLA/" + h)
                    ws_HLA = wb.create_sheet(h.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_HLA, content)
    
    if "TMB" in dir_list:
        tmb = os.listdir(resultsDir + "/TMB")
        if len(tmb) != 0:
            for t in tmb:
                if sampleID in t:
                    content = read_txt(resultsDir + "/TMB/" + t)
                    ws_TMB = wb.create_sheet(t.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_TMB, content)

    if "LOH" in dir_list:
        loh = os.listdir(resultsDir + "/LOH")
        if len(loh) != 0:
            for l in loh:
                if sampleID in l:
                    content = read_txt(resultsDir + "/LOH/" + l)
                    ws_LOH = wb.create_sheet(t.replace(".txt", "").replace(sampleID + ".", ""))
                    write_excel(ws_LOH, content)

    if "Neoantigen" in dir_list:
        neo = os.listdir(resultsDir + "/Neoantigen")
        if len(neo) != 0:
            for n in neo:
                if ".txt" in n:
                    if sampleID in n:
                        content = read_txt(resultsDir + "/Neoantigen/" + n)
                        ws_Neo = wb.create_sheet(n.replace(".filter.txt", "").replace(sampleID + ".", ""))
                        write_excel(ws_Neo, content)


    wb.remove(wb["Sheet"])
    try:
        wb.save(resultsDir + "/results/" + sampleID + ".xlsx")
    except:
        print("数据表不足保存")
    print(sampleID + " 结果已汇总到excel表格中： " + resultsDir + "/results/" + sampleID + ".xlsx")