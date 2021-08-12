# coding=utf-8
# pzw
# 20210812

import os
import openpyxl


reportList = os.listdir("Report")

bacteriaList = []
virusList = []
fungiList = []
parasiteList = []
for i in reportList:
    if "xlsx" in i:
        if not i.startswith("~"):
            book = openpyxl.load_workbook("Report/" + i)
            bacteria = book["bacteria"]
            virus = book["virus"]
            fungi = book["fungi"]
            parasite = book["parasite"]

            n = 2
            while bacteria.cell(n, 1).value != None:
                genus = bacteria.cell(n, 1).value
                genus_en = bacteria.cell(n, 2).value
                species = bacteria.cell(n, 6).value
                species_en = bacteria.cell(n, 7).value
                gram = bacteria.cell(n, 12).value
                bacteriaList.append("\t".join([genus, genus_en, species, species_en, gram]))
                n += 1
            
            j = 2
            while virus.cell(j, 1).value != None:
                genus = virus.cell(j, 1).value
                genus_en = virus.cell(j, 2).value
                species = virus.cell(j, 6).value
                species_en = virus.cell(j, 7).value
                virusList.append("\t".join([genus, genus_en, species, species_en]))
                j += 1

            k = 2
            while fungi.cell(k, 1).value != None:
                genus = fungi.cell(k, 1).value
                genus_en = fungi.cell(k, 2).value
                species = fungi.cell(k, 6).value
                species_en = fungi.cell(k, 7).value
                fungiList.append("\t".join([genus, genus_en, species, species_en]))
                k += 1

            m = 2
            while parasite.cell(m, 1).value != None:
                genus = parasite.cell(m, 1).value
                genus_en = parasite.cell(m, 2).value
                species = parasite.cell(m, 6).value
                species_en = parasite.cell(m, 7).value
                parasiteList.append("\t".join([genus, genus_en, species, species_en]))
                m += 1

            book.close()
            
bacteriaList = list(set(bacteriaList))
virusList = list(set(virusList))
fungiList = list(set(fungiList))
parasiteList = list(set(parasiteList))

bacteriaOutput = open("bacteria.txt", "w", encoding="utf-8")
virusOutput = open("virus.txt", "w", encoding="utf-8")
fungiOutput = open("fungi.txt", "w", encoding="utf-8")
parasiteOutput = open("parasite.txt", "w", encoding="utf-8")

bacteriaOutput.write("属\t属（英文名）\t种\t种（英文名）\t革兰氏染色\n")
virusOutput.write("属\t属（英文名）\t种\t种（英文名）\n")
fungiOutput.write("属\t属（英文名）\t种\t种（英文名）\n")
parasiteOutput.write("属\t属（英文名）\t种\t种（英文名）\n")

for n in bacteriaList:
    bacteriaOutput.write(n + "\n")
bacteriaOutput.close()
for j in virusList:
    virusOutput.write(j + "\n")
virusOutput.close()
for k in fungiList:
    fungiOutput.write(k + "\n")
fungiOutput.close()
for m in parasiteList:
    parasiteOutput.write(m + "\n")
parasiteOutput.close()
